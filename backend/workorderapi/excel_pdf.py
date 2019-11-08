from subprocess import call
from shutil import copyfile
from openpyxl import load_workbook
from pathlib import Path
from tempfile import mkdtemp
import os
import subprocess


class Spreadsheet():
    # takes a Path object and redsis-job/None
    def __init__(self, file, job):
        self.file = file
        self.job = job
        self.commercial = False
        self.start = 0
        self.end = 0
        self.bottom = 0

    def set_status(self, message):
        if self.job:
            text = self.job.meta['status']
            self.job.meta['status'] = text + message + '\n'
            self.job.save_meta()
        else:
            print(message)

    def prep_sheet(self):
        self.folder = mkdtemp()
        if self.file.suffix.lower() == ".xls":
            self.set_status('Converting From xls To xlsx')

            call(['/usr/local/bin/unoconv',
                  '-f', 'xlsx',
                  '--output=' + self.folder + '/workorder.xlsx',
                  self.file.as_posix()])

            self.set_status('Sheet Converted To xlsx')
        elif self.file.suffix == ".xlsx":
            self.set_status('Copying File')
            copyfile(self.file.as_posix(), self.folder + '/workorder.xlsx')
            self.set_status('File Copied')

    def opensheet(self):
        self.wb = load_workbook(self.folder + '/workorder.xlsx')
        self.ws = self.wb.worksheets[0]

    def find_edges(self):
        self.set_status('Edge Detction In Progress')
        for i in range(self.ws.max_row, 0, -1):
            v1 = self.ws.cell(column=1, row=i).value
            v2 = self.ws.cell(column=2, row=i).value
            p1 = self.ws.cell(column=1, row=i).fill.patternType

            # we hit a cell with a fill pattern or content
            if not self.bottom and (p1 or v1):
                self.bottom = i
                continue

            # bottom of search area + 1
            if v1 == 'PAINT OPTIONS':
                self.end = i
                self.commercial = True
                continue
            if v1 == 'PAINT' or v1 == 'PAINT / VINYL':
                self.end = i
                continue

            # top of search area
            if v2 == 'ENGINE OUTFITTING OPTIONS':
                self.start = i
                self.set_status('Edges Detected')
                break
            if v1 == 'ENGINE OPTIONS':
                self.start = i
                self.set_status('Edges Detected')
                break

    def create_text_version(self):
        self.set_status('Compiling Line Items')
        self.text = []
        for i in range(self.start, self.end):
            value = ""
            if self.ws.cell(column=1, row=i).value:
                value += str(self.ws.cell(column=1, row=i).value)
            if self.ws.cell(column=2, row=i).value:
                value += ' ' + str(self.ws.cell(column=2, row=i).value)
            if self.ws.cell(column=9, row=i).value:
                value += ' ' + str(self.ws.cell(column=9, row=i).value)
            self.text.append(value)

    def save_text_version(self):
        with open('/opt/workorders_git/' + self.file.stem + '.txt',
                  'w') as file:
            for text in self.text:
                file.write(text + "\n")

    def crop_sheet(self):
        self.set_status('Cropping Sheet')
        for sheet in self.wb.get_sheet_names()[1:]:
            ws = self.wb.get_sheet_by_name(sheet)
            self.wb.remove_sheet(ws)
        if self.ws.max_column > 9:
            self.ws.delete_cols(10, self.ws.max_column - 8)
        if self.ws.max_row > self.bottom:
            self.ws.delete_rows(self.bottom + 1,
                                self.ws.max_row - self.bottom - 5)

    def set_print_properties(self):
        self.set_status('Setting Printer Properties')
        self.ws.print_area = 'A1:I' + str(self.bottom)
        self.ws.sheet_properties.pageSetUpPr.fitToPage = True
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = False

    def savesheet_as(self, file):
        self.set_status('Saving Sheet')
        self.wb.save(file)

    def delete_sheet(self):
        self.set_status('Removing Temporay Files')
        os.remove(self.folder + '/workorder.xlsx')
        os.rmdir(self.folder)

    def create_pdf(self):
        self.set_status('Creating PDF')
        subprocess.call(['/usr/local/bin/unoconv',
                         '-f',
                         'pdf',
                         '--output=' + '/samba/shares/production' +
                         '/OWNCLOUD/Work Orders/' +
                         self.file.name,
                         self.folder + '/workorder.xlsx'])

    def proccess_sheet(self):
        self.prep_sheet()
        self.opensheet()
        self.find_edges()
        self.create_text_version()
        self.crop_sheet()
        self.set_print_properties()
        self.save_text_version()
        self.wb.save(self.folder + '/workorder.xlsx')
        if self.file.parts[4].lower() == 'boats in production':
            self.create_pdf()
        else:
            self.set_status('Skipping PDF Creation')
        self.delete_sheet()
        self.set_status('Now Tracking {}'.format(self.file.name))


if __name__ == '__main__':
    """
    file = Path(
        "/samba/shares/production/Boats in Production/29123 A020 3Rivers - " +
        "Maschmedt 2020 29' Seahawk Walkaround #OS4 9-12-19/Work Order/" +
        "29123 A020 - 3Rivers - Maschedt - 2020 29' Seahawk OS 2900SXL " +
        "Walkaround #OS4  09-12-19 WORK ORDER.xls"
    )
    file = Path("/samba/shares/production/Boats in Production/34277 G919 " +
                "Huntington NY 34'x12' Liberty (4-16-19)/A 1 Workorder/" +
                "Huntington WO .xlsx")
    """
    file = Path("/samba/shares/production/Boats in Production/21101 K920 " +
                "Clemens EUG 2020 21' Seahawk OB #12 9-4-19/Work Order/" +
                "21101 K920- Clemens EUG - 2020 21' Seahawk OB #12  " +
                "MED-SDB-TD  09-04-19 WORK ORDER.xls")

    sheet = Spreadsheet(file, None)
    sheet.proccess_sheet()
